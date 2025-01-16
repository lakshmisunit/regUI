#!/usr/bin/env python3
from openpyxl import load_workbook
from reg_classes import Reg, RegBlock
from regCodeGen import processRegBlock, processRegMacros, processRegMacros_D
from argParser import argParser
import config
import sys
import os, shutil

class xl_reader:
    def __init__(self):
        self.args = argParser().getArgs()
        self.dest_filename = self.args.input
        self.block = self.args.block
        self.block_addr = self.args.block_addr
        self.wb = load_workbook(filename=self.dest_filename)
        self.output_path = self.get_output_path('output')#'/home/projects/sw_model/lakshmim/RTL_UI/output'
        self.mname = self.get_config_value('mname').format(self.block)
        self.fname = self.get_config_value('fname').format(self.block)
        self.dname = self.get_config_value('dname').format(self.block)
        self.cname = self.get_config_value('cname').format(self.block)
        self.x, self.y = None, None
        self.entries, self.macros = [], []
        self.process_excel_file()
        self.write_files()
    
    def get_output_path(self, name):
        return getattr(self.args, name, None) or '/home/projects/sw_model/lakshmim/RTL_UI/output'

    def get_config_value(self, name):
        return getattr(self.args, name, None) or getattr(config, name)

    def process_excel_file(self):
        #Read the Excel sheet and process the register blocks.
        ws = self.wb[self.block]
        state = 'IDLE'
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
            if state == 'IDLE':
                self.process_idle_state(row)
                state = 'BLOCK'
            elif state == 'BLOCK':
                self.process_block_state(row)
                state = 'REG'
            elif state == 'REG':
                self.process_reg_state(row)
                #print(f"{row} = {row[0]}\n")
                if(row[0] != None):
                    state = 'BLOCK' if row[0] != 'END' else ''

    def process_idle_state(self, row):
        # process the idle state when a new block starts.
        if row[0] and row[0] != 'END':
            self.macros.append(f'`define {row[0]}__BASE_ADDR {row[1]}')
            self.x = RegBlock(row[0], row[1], row[9])
    
    def process_block_state(self, row):
        # process the block state when we are working with register block headers.
        if row[2]:
            self.macros.append(f'`define {row[2]}__REG_OFFSET {row[3]}')
            if self.y:
                self.x.insertReg(self.y)
            self.y = Reg(int(self.block_addr), row[2], row[7], row[3], row[5], row[9])

    def process_reg_state(self, row):
        # Process the register fields and update the register block.
        if row[4] is not None:
            self.y.insertFld(row[4], row[5], row[6], row[7], row[8], row[9])
        elif row[2]:
            self.macros.append(f'`define {row[2]}__REG_OFFSET {row[3]}')
            self.x.insertReg(self.y)
            self.y = Reg(int(self.block_addr), row[2], row[7], row[3], row[5], row[9])
            self.y.fields = []
        
        elif row[0]:
            self.x.insertReg(self.y)
            self.entries.append(self.x)
            if row[0] != 'END':
                self.x = RegBlock(row[0], row[1], row[2])
                self.y = None
            if (row[0] == 'END'):
                self.entries.append(self.x)


    def write_files(self):
        #Write the macros and register blocks to their respective files
        self.x.setMacros(self.macros)
    
         # Write to the files using the appropriate processing function
        self.write_to_file(self.fname, processRegBlock)
        self.write_to_file(self.dname, processRegMacros)
        self.write_to_file(self.cname, processRegMacros_D)
        self.move_files()

    def move_files(self):
        cwd = os.getcwd()
        shutil.move(os.path.join(cwd, self.fname), os.path.join(self.output_path, self.fname))
        shutil.move(os.path.join(cwd, self.dname), os.path.join(self.output_path, self.dname))
        shutil.move(os.path.join(cwd, self.cname), os.path.join(self.output_path, self.cname))
        shutil.copy(os.path.join(cwd, 'regLibCells.v'), os.path.join(self.output_path, 'regLibCells.v'))

    
    def write_to_file(self, filename, processing_function):
        #function to handle writing to a file
        with open(filename, 'w') as f:
            if processing_function == processRegBlock:
                processing_function(f, self.mname, self.entries[0].regArray)
            else:
                processing_function(f, self.entries[0].regArray)

        
            #processing_function(f, self.mname, self.entries[0].regArray if processing_function == processRegBlock else self.entries[0].regArray)
    

if __name__ == "__main__":
    #usage ./xl_reader.py --dest_filename ../integration_tool/ral_code_gen/i_uart_registers.xlsx --block uart_address_block --block_addr 000
    xl_reader()


